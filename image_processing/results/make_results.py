# Script used to create the result plots as well as formatted excel workbooks for quick results interpretation
import csv, os
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side, BORDER_MEDIUM
from openpyxl.formatting.rule import Rule, FormatObject
from openpyxl.formatting.rule import CellIsRule
from openpyxl.formatting.rule import ColorScale, Color
import matplotlib.pyplot as plt
from matplotlib import rcParams

from image_processing.constants import ROOT_DIR


CORRECT_RESULTS_LOC = ROOT_DIR + "\\data\\correct_results.csv"


def make_csv(data_loc, dir_name, file_name, flag_templates, flag_knn):
    """Makes a csv of containing all the overall metrics for each classification run"""

    correct_results = pd.read_csv(CORRECT_RESULTS_LOC)
    images = np.array(correct_results['image'])

    tc1_correct_result = correct_results[['image', 'tc1']]
    tc2_correct_result = correct_results[['image', 'tc2']]
    tc3_correct_result = correct_results[['image', 'tc3']]
    tc4_correct_result = correct_results[['image', 'tc4']]
    tc5_correct_result = correct_results[['image', 'tc5']]
    tc6_correct_result = correct_results[['image', 'tc6']]

    if not os.path.exists(dir_name):
        os.makedirs(dir_name)

    if os.path.exists(dir_name + file_name):
        os.remove(dir_name + file_name)

    with open(dir_name + file_name, 'a+') as csv_file:
        writer = csv.writer(csv_file, delimiter=',', quotechar='|', quoting=csv.QUOTE_MINIMAL,
                            lineterminator='\n')
        row = [
            "Fourier Descriptors",
            "Neighbors",
            "Accuracy",
            "Average Reliability Correct",
            "Standard Deviation Reliability Correct",
            "Average Euclidean Distance Correct",
            "Standard Deviation Euclidean Distance Correct",
            "Average Reliability Incorrect",
            "Standard Deviation Reliability Incorrect",
            "Average Euclidean Distance Incorrect",
            "Standard Deviation Euclidean Distance Incorrect",
            "undetectable"
        ]

        writer.writerow(row)
        csv_file.close()

    fourier_descriptors = [10, 20, 30, 40, 50]
    neighbours = [5, 7, 9, 11]

    n_5_correct_rel = []
    n_7_correct_rel = []
    n_9_correct_rel = []
    n_11_correct_rel = []

    n_5_incorrect_rel = []
    n_7_incorrect_rel = []
    n_9_incorrect_rel = []
    n_11_incorrect_rel = []

    n_5_correct_dist = []
    n_7_correct_dist = []
    n_9_correct_dist = []
    n_11_correct_dist = []

    n_5_incorrect_dist = []
    n_7_incorrect_dist = []
    n_9_incorrect_dist = []
    n_11_incorrect_dist = []

    for i in fourier_descriptors:
        for j in neighbours:
            data_loc_final = data_loc + '%d_fourier_descriptors\\%d_nn\\results.csv' % (i, j)
            data = pd.read_csv(data_loc_final)

            tc1_result = data[['Image', 'Fourier Descriptors', 'Neighbors', 'TC 1 result', 'TC 1 rel', 'TC 1 dist']]
            tc2_result = data[['Image', 'Fourier Descriptors', 'Neighbors', 'TC 2 result', 'TC 2 rel', 'TC 2 dist']]
            tc3_result = data[['Image', 'Fourier Descriptors', 'Neighbors', 'TC 3 result', 'TC 3 rel', 'TC 3 dist']]
            tc4_result = data[['Image', 'Fourier Descriptors', 'Neighbors', 'TC 4 result', 'TC 4 rel', 'TC 4 dist']]
            tc5_result = data[['Image', 'Fourier Descriptors', 'Neighbors', 'TC 5 result', 'TC 5 rel', 'TC 5 dist']]
            tc6_result = data[['Image', 'Fourier Descriptors', 'Neighbors', 'TC 6 result', 'TC 6 rel', 'TC 6 dist']]

            correct = 0
            incorrect = 0
            undetectable = 0

            correct_rel = []
            incorrect_rel = []

            correct_dist = []
            incorrect_dist = []

            for image in images:
                tc1 = tc1_result.loc[tc1_result['Image'] == image]
                tc2 = tc2_result.loc[tc2_result['Image'] == image]
                tc3 = tc3_result.loc[tc3_result['Image'] == image]
                tc4 = tc4_result.loc[tc4_result['Image'] == image]
                tc5 = tc5_result.loc[tc5_result['Image'] == image]
                tc6 = tc6_result.loc[tc6_result['Image'] == image]

                tc = [tc1, tc2, tc3, tc4, tc5, tc6]

                image_tc1_correct_results = tc1_correct_result.loc[tc1_correct_result['image'] == image]
                image_tc2_correct_results = tc2_correct_result.loc[tc2_correct_result['image'] == image]
                image_tc3_correct_results = tc3_correct_result.loc[tc3_correct_result['image'] == image]
                image_tc4_correct_results = tc4_correct_result.loc[tc4_correct_result['image'] == image]
                image_tc5_correct_results = tc5_correct_result.loc[tc5_correct_result['image'] == image]
                image_tc6_correct_results = tc6_correct_result.loc[tc6_correct_result['image'] == image]

                results = [
                    image_tc1_correct_results,
                    image_tc2_correct_results,
                    image_tc3_correct_results,
                    image_tc4_correct_results,
                    image_tc5_correct_results,
                    image_tc6_correct_results
                ]

                for ind, el in enumerate(tc):

                    correct_digit = int(results[ind]['tc%d' % (ind + 1)])
                    predicted_digit = int(el['TC %d result' % (ind + 1)])

                    if el['TC %d result' % (ind + 1)].isnull().values.any():
                        undetectable += 1
                    elif correct_digit == predicted_digit:
                        correct += 1
                        correct_rel.append(el['TC %d rel' % (ind + 1)].iloc[0])
                        correct_dist.append(el['TC %d dist' % (ind + 1)].iloc[0])
                    else:
                        incorrect += 1
                        incorrect_rel.append(el['TC %d rel' % (ind + 1)].iloc[0])
                        incorrect_dist.append(el['TC %d dist' % (ind + 1)].iloc[0])

            with open(dir_name + file_name, 'a+') as csv_file:
                writer = csv.writer(csv_file, delimiter=',', quotechar='|', quoting=csv.QUOTE_MINIMAL,
                                    lineterminator='\n')

                row = [
                    i,
                    j,
                    "%.1f" % ((correct / ((len(images) * 6) - undetectable)) * 100),
                    "%.2f" % np.mean(correct_rel),
                    "%.2f" % np.std(correct_rel),
                    "%.2f" % np.mean(correct_dist),
                    "%.2f" % np.std(correct_dist),
                    "%.2f" % np.mean(incorrect_rel),
                    "%.2f" % np.std(incorrect_rel),
                    "%.2f" % np.mean(incorrect_dist),
                    "%.2f" % np.std(incorrect_dist),
                    undetectable
                ]

                writer.writerow(row)
                csv_file.close()

                if j == 5:
                    n_5_correct_rel.append(correct_rel)
                    n_5_incorrect_rel.append(incorrect_rel)
                    n_5_correct_dist.append(correct_dist)
                    n_5_incorrect_dist.append(incorrect_dist)

                if j == 7:
                    n_7_correct_rel.append(correct_rel)
                    n_7_incorrect_rel.append(incorrect_rel)
                    n_7_correct_dist.append(correct_dist)
                    n_7_incorrect_dist.append(incorrect_dist)

                if j == 9:
                    n_9_correct_rel.append(correct_rel)
                    n_9_incorrect_rel.append(incorrect_rel)
                    n_9_correct_dist.append(correct_dist)
                    n_9_incorrect_dist.append(incorrect_dist)

                if j == 11:
                    n_11_correct_rel.append(correct_rel)
                    n_11_incorrect_rel.append(incorrect_rel)
                    n_11_correct_dist.append(correct_dist)
                    n_11_incorrect_dist.append(incorrect_dist)

    if not os.path.exists(ROOT_DIR + "\\results\\digit_classification\\box_plots\\"):
        os.makedirs(ROOT_DIR + "\\results\\digit_classification\\box_plots\\structured\\uniform")
        os.makedirs(ROOT_DIR + "\\results\\digit_classification\\box_plots\\structured\\weighted")
        os.makedirs(ROOT_DIR + "\\results\\digit_classification\\box_plots\\random\\uniform")
        os.makedirs(ROOT_DIR + "\\results\\digit_classification\\box_plots\\random\\weighted")

    from matplotlib.pylab import plot, show, savefig, xlim, figure, ylim, legend, boxplot, setp, axes

    fig, (ax2, ax3) = plt.subplots(nrows=1, ncols=2, sharex=True, sharey=True, figsize=(15, 8),
                                   gridspec_kw={'width_ratios': [3, 1]})

    rcParams.update({'font.size': 16})

    # function for setting the colors of the box plots pairs
    def setBoxColors(bp, lwidth):

        setp(bp['boxes'][0], color='#1f78b4')
        setp(bp['caps'][0], color='#ffffff')
        setp(bp['caps'][1], color='#ffffff')
        setp(bp['whiskers'][0], color='#ffffff')
        setp(bp['whiskers'][1], color='#ffffff')
        setp(bp['medians'][0], color='#1f78b4',  linewidth=lwidth)

        setp(bp['boxes'][1], color='#33a02c')
        setp(bp['caps'][2], color='#ffffff')
        setp(bp['caps'][3], color='#ffffff')
        setp(bp['whiskers'][2], color='#ffffff')
        setp(bp['whiskers'][3], color='#ffffff')
        setp(bp['medians'][1], color='#33a02c',  linewidth=lwidth)

        setp(bp['boxes'][2], color='#e31a1c')
        setp(bp['caps'][4], color='#ffffff')
        setp(bp['caps'][5], color='#ffffff')
        setp(bp['whiskers'][4], color='#ffffff')
        setp(bp['whiskers'][5], color='#ffffff')
        setp(bp['medians'][2], color='#e31a1c',  linewidth=lwidth)

        setp(bp['boxes'][3], color='#ff7f00')
        setp(bp['caps'][6], color='#ffffff')
        setp(bp['caps'][7], color='#ffffff')
        setp(bp['whiskers'][6], color='#ffffff')
        setp(bp['whiskers'][7], color='#ffffff')
        setp(bp['medians'][3], color='#ff7f00',  linewidth=lwidth)

        setp(bp['boxes'][4], color='#8856a7')
        setp(bp['caps'][8], color='#ffffff')
        setp(bp['caps'][9], color='#ffffff')
        setp(bp['whiskers'][8], color='#ffffff')
        setp(bp['whiskers'][9], color='#ffffff')
        setp(bp['medians'][4], color='#8856a7',  linewidth=lwidth)

    box_colors = ['#b3cde3', '#ccebc5', '#fbb4ae', '#fed9a6', '#decbe4']
    pd.set_option('display.width', 400)
    pd.set_option('display.max_columns', 10)

    print(file_name)

    stats_dir = ROOT_DIR + "\\results\\digit_classification\\statistics\\"

    if not os.path.exists(stats_dir):
        os.mkdir(stats_dir)

    if flag_templates == 0:
        dir_name_2 = ROOT_DIR + "\\results\\digit_classification\\statistics\\structured\\"

        if not os.path.exists(dir_name_2):
            os.mkdir(dir_name_2)

    else:
        dir_name_2 = ROOT_DIR + "\\results\\digit_classification\\statistics\\random\\"

        if not os.path.exists(dir_name_2):
            os.mkdir(dir_name_2)

    if flag_knn == 0:
        dir_name_2 += "uniform\\"

        if not os.path.exists(dir_name_2):
            os.mkdir(dir_name_2)

    else:
        dir_name_2 += "weighted\\"

        if not os.path.exists(dir_name_2):
            os.mkdir(dir_name_2)

    file_name_rel = "reliability_score.csv"

    if os.path.exists(dir_name_2+file_name_rel):
        os.remove(dir_name_2+file_name_rel)

    stat_names = ['mean', 'std', 'min', '25%', '50%', '75%', 'max']
    # Correct Reliability Score
    # 5 neighbours boxplot
    # bp = ax1.boxplot(n_5_correct_rel, positions=[1, 2, 3, 4, 5], widths=0.8, showfliers=False, patch_artist=True,
    #                  showcaps=True)
    # setBoxColors(bp)
    ldf_dic = {'10 fourier descriptors': pd.Series(n_5_correct_rel[0]),
               '20 fourier descriptors': pd.Series(n_5_correct_rel[1]),
               '30 fourier descriptors': pd.Series(n_5_correct_rel[2]),
               '40 fourier descriptors': pd.Series(n_5_correct_rel[3]),
               '50 fourier descriptors': pd.Series(n_5_correct_rel[4]),
               }
    ldf = pd.DataFrame(ldf_dic)
    stats = ldf.describe(include='all')
    print(stats)

    with open(dir_name_2 + file_name_rel, 'a+') as csv_file:
        writer = csv.writer(csv_file, delimiter=',', quotechar='|', quoting=csv.QUOTE_MINIMAL,
                            lineterminator='\n')

        writer.writerow(['Correct'])
        writer.writerow(['5 Neighbors'])
        writer.writerow(['','10 Fourier Descriptors','20 Fourier Descriptors','30 Fourier Descriptors','40 Fourier Descriptors','50 Fourier Descriptors'])

        for index, row in stats.iterrows():
            writer.writerow([index] + list(row.values.round(2)))
        csv_file.close()

    # for patch, color in zip(bp['boxes'], box_colors):
    #     patch.set_facecolor(color)

    # 7 neighbours boxplot
    # bp = ax1.boxplot(n_7_correct_rel, positions=[7, 8, 9, 10, 11], widths=0.8, showfliers=False, patch_artist=True,
    #                  showcaps=True)
    # setBoxColors(bp)
    ldf_dic = {'10 fourier descriptors': pd.Series(n_7_correct_rel[0]),
               '20 fourier descriptors': pd.Series(n_7_correct_rel[1]),
               '30 fourier descriptors': pd.Series(n_7_correct_rel[2]),
               '40 fourier descriptors': pd.Series(n_7_correct_rel[3]),
               '50 fourier descriptors': pd.Series(n_7_correct_rel[4]),
               }
    ldf = pd.DataFrame(ldf_dic)

    stats = ldf.describe(include='all')

    with open(dir_name_2 + file_name_rel, 'a+') as csv_file:
        writer = csv.writer(csv_file, delimiter=',', quotechar='|', quoting=csv.QUOTE_MINIMAL,
                            lineterminator='\n')
        writer.writerow(['7 Neighbors'])
        writer.writerow(['','10 Fourier Descriptors','20 Fourier Descriptors','30 Fourier Descriptors','40 Fourier Descriptors','50 Fourier Descriptors'])
        for index, row in stats.iterrows():
            writer.writerow([index] + list(row.values.round(2)))
        csv_file.close()
    # for patch, color in zip(bp['boxes'], box_colors):
    #     patch.set_facecolor(color)

    # 9 neighbours boxplot
    # bp = ax1.boxplot(n_9_correct_rel, positions=[13, 14, 15, 16, 17], widths=0.8, showfliers=False, patch_artist=True,
    #                  showcaps=True)
    # setBoxColors(bp)
    ldf_dic = {'10 fourier descriptors': pd.Series(n_9_correct_rel[0]),
               '20 fourier descriptors': pd.Series(n_9_correct_rel[1]),
               '30 fourier descriptors': pd.Series(n_9_correct_rel[2]),
               '40 fourier descriptors': pd.Series(n_9_correct_rel[3]),
               '50 fourier descriptors': pd.Series(n_9_correct_rel[4]),
               }
    ldf = pd.DataFrame(ldf_dic)
    stats = ldf.describe(include='all')

    with open(dir_name_2 + file_name_rel, 'a+') as csv_file:
        writer = csv.writer(csv_file, delimiter=',', quotechar='|', quoting=csv.QUOTE_MINIMAL,
                            lineterminator='\n')

        writer.writerow(['9 Neighbors'])
        writer.writerow(['','10 Fourier Descriptors','20 Fourier Descriptors','30 Fourier Descriptors','40 Fourier Descriptors','50 Fourier Descriptors'])

        for index, row in stats.iterrows():
            writer.writerow([index] + list(row.values.round(2)))
        csv_file.close()
    # for patch, color in zip(bp['boxes'], box_colors):
    #     patch.set_facecolor(color)

    # 11 neighbours boxplot
    # bp = ax1.boxplot(n_11_correct_rel, positions=[19, 20, 21, 22, 23], widths=0.8, showfliers=False, patch_artist=True,
    #                  showcaps=True)
    # setBoxColors(bp)
    ldf_dic = {'10 fourier descriptors': pd.Series(n_11_correct_rel[0]),
               '20 fourier descriptors': pd.Series(n_11_correct_rel[1]),
               '30 fourier descriptors': pd.Series(n_11_correct_rel[2]),
               '40 fourier descriptors': pd.Series(n_11_correct_rel[3]),
               '50 fourier descriptors': pd.Series(n_11_correct_rel[4]),
               }
    ldf = pd.DataFrame(ldf_dic)
    stats = ldf.describe(include='all')

    with open(dir_name_2 + file_name_rel, 'a+') as csv_file:
        writer = csv.writer(csv_file, delimiter=',', quotechar='|', quoting=csv.QUOTE_MINIMAL,
                            lineterminator='\n')

        writer.writerow(['11 Neighbors'])
        writer.writerow(['','10 Fourier Descriptors','20 Fourier Descriptors','30 Fourier Descriptors','40 Fourier Descriptors','50 Fourier Descriptors'])

        for index, row in stats.iterrows():
            writer.writerow([index] + list(row.values.round(2)))
        csv_file.close()
    # for patch, color in zip(bp['boxes'], box_colors):
    #     patch.set_facecolor(color)

    # Incorrect
    # 5 neighbours boxplot
    bp = ax2.boxplot(n_5_incorrect_rel, positions=[1, 2, 3, 4, 5], widths=0.8, showfliers=False, patch_artist=True,
                     showcaps=True)
    setBoxColors(bp, 2)
    ldf_dic = {'10 fourier descriptors': pd.Series(n_5_incorrect_rel[0]),
               '20 fourier descriptors': pd.Series(n_5_incorrect_rel[1]),
               '30 fourier descriptors': pd.Series(n_5_incorrect_rel[2]),
               '40 fourier descriptors': pd.Series(n_5_incorrect_rel[3]),
               '50 fourier descriptors': pd.Series(n_5_incorrect_rel[4]),
               }
    ldf = pd.DataFrame(ldf_dic)
    stats = ldf.describe(include='all')

    with open(dir_name_2 + file_name_rel, 'a+') as csv_file:
        writer = csv.writer(csv_file, delimiter=',', quotechar='|', quoting=csv.QUOTE_MINIMAL,
                            lineterminator='\n')

        writer.writerow(['Incorrect'])
        writer.writerow(['5 Neighbors'])
        writer.writerow(['','10 Fourier Descriptors','20 Fourier Descriptors','30 Fourier Descriptors','40 Fourier Descriptors','50 Fourier Descriptors'])

        for index, row in stats.iterrows():
            writer.writerow([index] + list(row.values.round(2)))
        csv_file.close()

    for patch, color in zip(bp['boxes'], box_colors):
        patch.set_facecolor(color)

    # 7 neighbours boxplot
    bp = ax2.boxplot(n_7_incorrect_rel, positions=[7, 8, 9, 10, 11], widths=0.8, showfliers=False, patch_artist=True,
                     showcaps=True)
    setBoxColors(bp, 2)
    ldf_dic = {'10 fourier descriptors': pd.Series(n_7_incorrect_rel[0]),
               '20 fourier descriptors': pd.Series(n_7_incorrect_rel[1]),
               '30 fourier descriptors': pd.Series(n_7_incorrect_rel[2]),
               '40 fourier descriptors': pd.Series(n_7_incorrect_rel[3]),
               '50 fourier descriptors': pd.Series(n_7_incorrect_rel[4]),
               }
    ldf = pd.DataFrame(ldf_dic)

    stats = ldf.describe(include='all')
    with open(dir_name_2 + file_name_rel, 'a+') as csv_file:
        writer = csv.writer(csv_file, delimiter=',', quotechar='|', quoting=csv.QUOTE_MINIMAL,
                            lineterminator='\n')

        writer.writerow(['7 Neighbors'])
        writer.writerow(['','10 Fourier Descriptors','20 Fourier Descriptors','30 Fourier Descriptors','40 Fourier Descriptors','50 Fourier Descriptors'])

        for index, row in stats.iterrows():
            writer.writerow([index] + list(row.values.round(2)))
        csv_file.close()

    for patch, color in zip(bp['boxes'], box_colors):
        patch.set_facecolor(color)

    # 9 neighbours boxplot
    bp = ax2.boxplot(n_9_incorrect_rel, positions=[13, 14, 15, 16, 17], widths=0.8, showfliers=False, patch_artist=True,
                     showcaps=True)
    setBoxColors(bp, 2)
    ldf_dic = {'10 fourier descriptors': pd.Series(n_9_incorrect_rel[0]),
               '20 fourier descriptors': pd.Series(n_9_incorrect_rel[1]),
               '30 fourier descriptors': pd.Series(n_9_incorrect_rel[2]),
               '40 fourier descriptors': pd.Series(n_9_incorrect_rel[3]),
               '50 fourier descriptors': pd.Series(n_9_incorrect_rel[4]),
               }
    ldf = pd.DataFrame(ldf_dic)

    stats = ldf.describe(include='all')
    with open(dir_name_2 + file_name_rel, 'a+') as csv_file:
        writer = csv.writer(csv_file, delimiter=',', quotechar='|', quoting=csv.QUOTE_MINIMAL,
                            lineterminator='\n')

        writer.writerow(['9 Neighbors'])
        writer.writerow(['','10 Fourier Descriptors','20 Fourier Descriptors','30 Fourier Descriptors','40 Fourier Descriptors','50 Fourier Descriptors'])

        for index, row in stats.iterrows():
            writer.writerow([index] + list(row.values.round(2)))
        csv_file.close()

    for patch, color in zip(bp['boxes'], box_colors):
        patch.set_facecolor(color)

    # 11 neighbours boxplot
    bp = ax2.boxplot(n_11_incorrect_rel, positions=[19, 20, 21, 22, 23], widths=0.8, showfliers=False,
                     patch_artist=True, showcaps=True)
    setBoxColors(bp, 2)
    ldf_dic = {'10 fourier descriptors': pd.Series(n_11_incorrect_rel[0]),
               '20 fourier descriptors': pd.Series(n_11_incorrect_rel[1]),
               '30 fourier descriptors': pd.Series(n_11_incorrect_rel[2]),
               '40 fourier descriptors': pd.Series(n_11_incorrect_rel[3]),
               '50 fourier descriptors': pd.Series(n_11_incorrect_rel[4]),
               }
    ldf = pd.DataFrame(ldf_dic)
    stats = ldf.describe(include='all')
    with open(dir_name_2 + file_name_rel, 'a+') as csv_file:
        writer = csv.writer(csv_file, delimiter=',', quotechar='|', quoting=csv.QUOTE_MINIMAL,
                            lineterminator='\n')

        writer.writerow(['11 Neighbors'])
        writer.writerow(['','10 Fourier Descriptors','20 Fourier Descriptors','30 Fourier Descriptors','40 Fourier Descriptors','50 Fourier Descriptors'])

        for index, row in stats.iterrows():
            writer.writerow([index] + list(row.values.round(2)))
        csv_file.close()
    for patch, color in zip(bp['boxes'], box_colors):
        patch.set_facecolor(color)

    # Dummy for for legend

    b3_1 = ax3.bar(1, height=1, label='10 Fourier descriptors', color='#b3cde3', edgecolor='#1f78b4')
    b3_2 = ax3.bar(1, height=1, label='20 Fourier descriptors', color='#ccebc5', edgecolor='#33a02c')
    b3_3 = ax3.bar(1, height=1, label='30 Fourier descriptors', color='#fbb4ae', edgecolor='#e31a1c')
    b3_4 = ax3.bar(1, height=1, label='40 Fourier descriptors', color='#fed9a6', edgecolor='#ff7f00')
    b3_5 = ax3.bar(1, height=1, label='50 Fourier descriptors', color='#decbe4', edgecolor='#8856a7')

    # set axes limits and labels
    xlim(0, 24)
    ylim(0, 1.03)
    # ax1.set_xticklabels(['5', '7', '9', '11'])
    ax2.set_xticklabels(['5', '7', '9', '11'], fontsize=12)
    ax2.tick_params(axis="y", labelsize=12)
    # ax1.set_xticks([3.5, 9.5, 15.5, 21.5])
    ax2.set_xticks([3, 9, 15, 21])
    # ax1.set_title("Correct", pad=15)
    # ax2.set_title("Incorrect", pad=15)
    # Add axis and chart labels
    ax2.set_xlabel('Number of Neighbors', labelpad=15, fontsize=16)
    ax2.set_ylabel('Reliability Score', labelpad=15, fontsize=16)

    ax2.axvline(x=6, linestyle=':', linewidth=.25)
    ax2.axvline(x=12, linestyle=':', linewidth=.25)
    ax2.axvline(x=18, linestyle=':', linewidth=.25)

    l = ax3.legend(loc="center", fontsize='large')
    # ax3.set_visible(False)

    for bar in ax3.patches:
        bar.set_visible(False)

    ax3.spines['top'].set_visible(False)
    ax3.spines['right'].set_visible(False)
    ax3.spines['left'].set_visible(False)
    ax3.spines['bottom'].set_visible(False)
    ax3.axis('off')

    if flag_templates is 0:
        if flag_knn is 0:
            plt.savefig(
                ROOT_DIR + "\\results\\digit_classification\\box_plots\\structured\\uniform\\rel_score_boxplot.pdf")
            plt.close()
        else:
            plt.savefig(
                ROOT_DIR + "\\results\\digit_classification\\box_plots\\structured\\weighted\\rel_score_boxplot.pdf")
            plt.close()
    else:
        if flag_knn is 0:
            plt.savefig(
                ROOT_DIR + "\\results\\digit_classification\\box_plots\\random\\uniform\\rel_score_boxplot.pdf")
            plt.close()
        else:
            plt.savefig(
                ROOT_DIR + "\\results\\digit_classification\\box_plots\\random\\weighted\\rel_score_boxplot.pdf")
            plt.close()

    fig, (ax1, ax2, ax3) = plt.subplots(nrows=1, ncols=3, sharex=True, sharey=True, figsize=(15, 8),
                                        gridspec_kw={'width_ratios': [3, 3, 1]})

    file_name_dist = "average_euclidean_distance.csv"

    rcParams.update({'font.size': 14})
    # Correct Avg. Euclidean Distance
    # 5 neighbours boxplot
    bp = ax1.boxplot(n_5_correct_dist, positions=[1, 2, 3, 4, 5], widths=0.8, showfliers=False, patch_artist=True,
                     showcaps=True)
    setBoxColors(bp, 1)
    ldf_dic = {'10 fourier descriptors': pd.Series(n_5_correct_dist[0]),
               '20 fourier descriptors': pd.Series(n_5_correct_dist[1]),
               '30 fourier descriptors': pd.Series(n_5_correct_dist[2]),
               '40 fourier descriptors': pd.Series(n_5_correct_dist[3]),
               '50 fourier descriptors': pd.Series(n_5_correct_dist[4]),
               }
    ldf = pd.DataFrame(ldf_dic)
    stats = ldf.describe(include='all')
    with open(dir_name_2 + file_name_dist, 'a+') as csv_file:
        writer = csv.writer(csv_file, delimiter=',', quotechar='|', quoting=csv.QUOTE_MINIMAL,
                            lineterminator='\n')

        writer.writerow(['Correct'])
        writer.writerow(['5 Neighbors'])
        writer.writerow(['','10 Fourier Descriptors','20 Fourier Descriptors','30 Fourier Descriptors','40 Fourier Descriptors','50 Fourier Descriptors'])

        for index, row in stats.iterrows():
            writer.writerow([index] + list(row.values.round(2)))
        csv_file.close()
    for patch, color in zip(bp['boxes'], box_colors):
        patch.set_facecolor(color)

    # 7 neighbours boxplot
    bp = ax1.boxplot(n_7_correct_dist, positions=[7, 8, 9, 10, 11], widths=0.8, showfliers=False, patch_artist=True,
                     showcaps=True)
    setBoxColors(bp, 1)
    ldf_dic = {'10 fourier descriptors': pd.Series(n_7_correct_dist[0]),
               '20 fourier descriptors': pd.Series(n_7_correct_dist[1]),
               '30 fourier descriptors': pd.Series(n_7_correct_dist[2]),
               '40 fourier descriptors': pd.Series(n_7_correct_dist[3]),
               '50 fourier descriptors': pd.Series(n_7_correct_dist[4]),
               }
    ldf = pd.DataFrame(ldf_dic)
    stats = ldf.describe(include='all')
    with open(dir_name_2 + file_name_dist, 'a+') as csv_file:
        writer = csv.writer(csv_file, delimiter=',', quotechar='|', quoting=csv.QUOTE_MINIMAL,
                            lineterminator='\n')

        writer.writerow(['7 Neighbors'])
        writer.writerow(['','10 Fourier Descriptors','20 Fourier Descriptors','30 Fourier Descriptors','40 Fourier Descriptors','50 Fourier Descriptors'])

        for index, row in stats.iterrows():
            writer.writerow([index] + list(row.values.round(2)))
        csv_file.close()

    for patch, color in zip(bp['boxes'], box_colors):
        patch.set_facecolor(color)

    # 9 neighbours boxplot
    bp = ax1.boxplot(n_9_correct_dist, positions=[13, 14, 15, 16, 17], widths=0.8, showfliers=False, patch_artist=True,
                     showcaps=True)
    setBoxColors(bp, 1)
    ldf_dic = {'10 fourier descriptors': pd.Series(n_9_correct_dist[0]),
               '20 fourier descriptors': pd.Series(n_9_correct_dist[1]),
               '30 fourier descriptors': pd.Series(n_9_correct_dist[2]),
               '40 fourier descriptors': pd.Series(n_9_correct_dist[3]),
               '50 fourier descriptors': pd.Series(n_9_correct_dist[4]),
               }
    ldf = pd.DataFrame(ldf_dic)
    stats = ldf.describe(include='all')
    with open(dir_name_2 + file_name_dist, 'a+') as csv_file:
        writer = csv.writer(csv_file, delimiter=',', quotechar='|', quoting=csv.QUOTE_MINIMAL,
                            lineterminator='\n')

        writer.writerow(['9 Neighbors'])
        writer.writerow(['','10 Fourier Descriptors','20 Fourier Descriptors','30 Fourier Descriptors','40 Fourier Descriptors','50 Fourier Descriptors'])

        for index, row in stats.iterrows():
            writer.writerow([index] + list(row.values.round(2)))
        csv_file.close()

    for patch, color in zip(bp['boxes'], box_colors):
        patch.set_facecolor(color)

    # 11 neighbours boxplot
    bp = ax1.boxplot(n_11_correct_dist, positions=[19, 20, 21, 22, 23], widths=0.8, showfliers=False, patch_artist=True,
                     showcaps=True)
    setBoxColors(bp, 1)
    ldf_dic = {'10 fourier descriptors': pd.Series(n_11_correct_dist[0]),
               '20 fourier descriptors': pd.Series(n_11_correct_dist[1]),
               '30 fourier descriptors': pd.Series(n_11_correct_dist[2]),
               '40 fourier descriptors': pd.Series(n_11_correct_dist[3]),
               '50 fourier descriptors': pd.Series(n_11_correct_dist[4]),
               }
    ldf = pd.DataFrame(ldf_dic)
    stats = ldf.describe(include='all')

    with open(dir_name_2 + file_name_dist, 'a+') as csv_file:
        writer = csv.writer(csv_file, delimiter=',', quotechar='|', quoting=csv.QUOTE_MINIMAL,
                            lineterminator='\n')

        writer.writerow(['11 Neighbors'])
        writer.writerow(['','10 Fourier Descriptors','20 Fourier Descriptors','30 Fourier Descriptors','40 Fourier Descriptors','50 Fourier Descriptors'])

        for index, row in stats.iterrows():
            writer.writerow([index] + list(row.values.round(2)))
        csv_file.close()

    for patch, color in zip(bp['boxes'], box_colors):
        patch.set_facecolor(color)

    # Incorrect
    # 5 neighbours boxplot
    bp = ax2.boxplot(n_5_incorrect_dist, positions=[1, 2, 3, 4, 5], widths=0.8, showfliers=False, patch_artist=True,
                     showcaps=True)
    setBoxColors(bp, 1)
    ldf_dic = {'10 fourier descriptors': pd.Series(n_5_incorrect_dist[0]),
               '20 fourier descriptors': pd.Series(n_5_incorrect_dist[1]),
               '30 fourier descriptors': pd.Series(n_5_incorrect_dist[2]),
               '40 fourier descriptors': pd.Series(n_5_incorrect_dist[3]),
               '50 fourier descriptors': pd.Series(n_5_incorrect_dist[4]),
               }
    ldf = pd.DataFrame(ldf_dic)
    stats = ldf.describe(include='all')
    with open(dir_name_2 + file_name_dist, 'a+') as csv_file:
        writer = csv.writer(csv_file, delimiter=',', quotechar='|', quoting=csv.QUOTE_MINIMAL,
                            lineterminator='\n')

        writer.writerow(['Incorrect'])
        writer.writerow(['5 Neighbors'])
        writer.writerow(['','10 Fourier Descriptors','20 Fourier Descriptors','30 Fourier Descriptors','40 Fourier Descriptors','50 Fourier Descriptors'])

        for index, row in stats.iterrows():
            writer.writerow([index] + list(row.values.round(2)))
        csv_file.close()

    for patch, color in zip(bp['boxes'], box_colors):
        patch.set_facecolor(color)

    # 7 neighbours boxplot
    bp = ax2.boxplot(n_7_incorrect_dist, positions=[7, 8, 9, 10, 11], widths=0.8, showfliers=False, patch_artist=True,
                     showcaps=True)
    setBoxColors(bp, 1)
    ldf_dic = {'10 fourier descriptors': pd.Series(n_7_incorrect_dist[0]),
               '20 fourier descriptors': pd.Series(n_7_incorrect_dist[1]),
               '30 fourier descriptors': pd.Series(n_7_incorrect_dist[2]),
               '40 fourier descriptors': pd.Series(n_7_incorrect_dist[3]),
               '50 fourier descriptors': pd.Series(n_7_incorrect_dist[4]),
               }
    ldf = pd.DataFrame(ldf_dic)
    stats = ldf.describe(include='all')
    with open(dir_name_2 + file_name_dist, 'a+') as csv_file:
        writer = csv.writer(csv_file, delimiter=',', quotechar='|', quoting=csv.QUOTE_MINIMAL,
                            lineterminator='\n')

        writer.writerow(['7 Neighbors'])
        writer.writerow(['','10 Fourier Descriptors','20 Fourier Descriptors','30 Fourier Descriptors','40 Fourier Descriptors','50 Fourier Descriptors'])

        for index, row in stats.iterrows():
            writer.writerow([index] + list(row.values.round(2)))
        csv_file.close()
    for patch, color in zip(bp['boxes'], box_colors):
        patch.set_facecolor(color)

    # 9 neighbours boxplot
    bp = ax2.boxplot(n_9_incorrect_dist, positions=[13, 14, 15, 16, 17], widths=0.8, showfliers=False,
                     patch_artist=True, showcaps=True)
    setBoxColors(bp, 1)
    ldf_dic = {'10 fourier descriptors': pd.Series(n_9_incorrect_dist[0]),
               '20 fourier descriptors': pd.Series(n_9_incorrect_dist[1]),
               '30 fourier descriptors': pd.Series(n_9_incorrect_dist[2]),
               '40 fourier descriptors': pd.Series(n_9_incorrect_dist[3]),
               '50 fourier descriptors': pd.Series(n_9_incorrect_dist[4]),
               }
    ldf = pd.DataFrame(ldf_dic)
    stats = ldf.describe(include='all')
    with open(dir_name_2 + file_name_dist, 'a+') as csv_file:
        writer = csv.writer(csv_file, delimiter=',', quotechar='|', quoting=csv.QUOTE_MINIMAL,
                            lineterminator='\n')

        writer.writerow(['9 Neighbors'])
        writer.writerow(['','10 Fourier Descriptors','20 Fourier Descriptors','30 Fourier Descriptors','40 Fourier Descriptors','50 Fourier Descriptors'])

        for index, row in stats.iterrows():
            writer.writerow([index] + list(row.values.round(2)))
        csv_file.close()

    for patch, color in zip(bp['boxes'], box_colors):
        patch.set_facecolor(color)

    # 11 neighbours boxplot
    bp = ax2.boxplot(n_11_incorrect_dist, positions=[19, 20, 21, 22, 23], widths=0.8, showfliers=False,
                     patch_artist=True, showcaps=True)
    setBoxColors(bp, 1)
    ldf_dic = {'10 fourier descriptors': pd.Series(n_11_incorrect_dist[0]),
               '20 fourier descriptors': pd.Series(n_11_incorrect_dist[1]),
               '30 fourier descriptors': pd.Series(n_11_incorrect_dist[2]),
               '40 fourier descriptors': pd.Series(n_11_incorrect_dist[3]),
               '50 fourier descriptors': pd.Series(n_11_incorrect_dist[4]),
               }
    ldf = pd.DataFrame(ldf_dic)
    stats = ldf.describe(include='all')
    with open(dir_name_2 + file_name_dist, 'a+') as csv_file:
        writer = csv.writer(csv_file, delimiter=',', quotechar='|', quoting=csv.QUOTE_MINIMAL,
                            lineterminator='\n')

        writer.writerow(['11 Neighbors'])
        writer.writerow(['','10 Fourier Descriptors','20 Fourier Descriptors','30 Fourier Descriptors','40 Fourier Descriptors','50 Fourier Descriptors'])

        for index, row in stats.iterrows():
            writer.writerow([index] + list(row.values.round(2)))
        csv_file.close()
    for patch, color in zip(bp['boxes'], box_colors):
        patch.set_facecolor(color)

    # Dummy for for legend

    b3_1 = ax3.bar(1, height=1, label='10 Fourier descriptors', color='#b3cde3', edgecolor='#1f78b4')
    b3_2 = ax3.bar(1, height=1, label='20 Fourier descriptors', color='#ccebc5', edgecolor='#33a02c')
    b3_3 = ax3.bar(1, height=1, label='30 Fourier descriptors', color='#fbb4ae', edgecolor='#e31a1c')
    b3_4 = ax3.bar(1, height=1, label='40 Fourier descriptors', color='#fed9a6', edgecolor='#ff7f00')
    b3_5 = ax3.bar(1, height=1, label='50 Fourier descriptors', color='#decbe4', edgecolor='#8856a7')

    # set axes limits and labels
    xlim(0, 24)
    ylim(0, 1)
    ax1.set_xticklabels(['5', '7', '9', '11'], fontsize=10)
    ax2.set_xticklabels(['5', '7', '9', '11'], fontsize=10)
    ax1.set_xticks([3, 9, 15, 21])
    ax2.set_xticks([3, 9, 15, 21])
    ax1.set_title("Correct", pad=15)
    ax2.set_title("Incorrect", pad=15)
    ax1.tick_params(axis="y", labelsize=10)
    ax2.tick_params(axis="y", labelsize=10)
    ax1.axvline(x=6, linestyle=':', linewidth=.25)
    ax1.axvline(x=12, linestyle=':', linewidth=.25)
    ax1.axvline(x=18, linestyle=':', linewidth=.25)
    ax2.axvline(x=6, linestyle=':', linewidth=.25)
    ax2.axvline(x=12, linestyle=':', linewidth=.25)
    ax2.axvline(x=18, linestyle=':', linewidth=.25)


    ax1.set_xlabel('Number of Neighbors', labelpad=15, fontsize=14)
    ax1.set_ylabel('Average Euclidean Distance in Feature Space', labelpad=15, fontsize=14)
    ax2.set_xlabel('Number of Neighbors', labelpad=15, fontsize=14)

    l = ax3.legend(loc="center", prop={'size': 12})
    # ax3.set_visible(False)

    for bar in ax3.patches:
        bar.set_visible(False)

    ax3.spines['top'].set_visible(False)
    ax3.spines['right'].set_visible(False)
    ax3.spines['left'].set_visible(False)
    ax3.spines['bottom'].set_visible(False)
    ax3.axis('off')

    if flag_templates is 0:
        if flag_knn is 0:
            plt.savefig(
                ROOT_DIR + "\\results\\digit_classification\\box_plots\\structured\\uniform\\euc_dist_boxplot.pdf")
            plt.close()
        else:
            plt.savefig(
                ROOT_DIR + "\\results\\digit_classification\\box_plots\\structured\\weighted\\euc_dist_boxplot.pdf")
            plt.close()
    else:
        if flag_knn is 0:
            plt.savefig(
                ROOT_DIR + "\\results\\digit_classification\\box_plots\\random\\uniform\\euc_dist_boxplot.pdf")
            plt.close()
        else:
            plt.savefig(
                ROOT_DIR + "\\results\\digit_classification\\box_plots\\random\\weighted\\euc_dist_boxplot.pdf")
            plt.close()


def make_csv_detailed(data_loc, dir_name, file_name, xlsx_name):
    """Makes a csv file containing the results for the digit classification for each tally counter"""
    correct_results = pd.read_csv(CORRECT_RESULTS_LOC)
    images = np.array(correct_results['image'])

    tc1_correct_result = correct_results[['image', 'tc1']]
    tc2_correct_result = correct_results[['image', 'tc2']]
    tc3_correct_result = correct_results[['image', 'tc3']]
    tc4_correct_result = correct_results[['image', 'tc4']]
    tc5_correct_result = correct_results[['image', 'tc5']]
    tc6_correct_result = correct_results[['image', 'tc6']]

    correct_results_list = [
        tc1_correct_result,
        tc2_correct_result,
        tc3_correct_result,
        tc4_correct_result,
        tc5_correct_result,
        tc6_correct_result,
    ]

    if not os.path.exists(dir_name):
        os.makedirs(dir_name)

    if os.path.exists(dir_name + file_name):
        os.remove(dir_name + file_name)

    if not os.path.exists(dir_name + file_name):
        with open(dir_name + file_name, 'a+') as csv_file:
            writer = csv.writer(csv_file, delimiter=',', quotechar='|', quoting=csv.QUOTE_MINIMAL, lineterminator='\n')
            row = [
                'image',
                'Fourier Descriptors',
                'Neighbors',
                'TC 1 result',
                'TC 2 result',
                'TC 3 result',
                'TC 4 result',
                'TC 5 result',
                'TC 6 result',
                'TC 1 rel',
                'TC 2 rel',
                'TC 3 rel',
                'TC 4 rel',
                'TC 5 rel',
                'TC 6 rel',
                'TC 1 dist',
                'TC 2 dist',
                'TC 3 dist',
                'TC 4 dist',
                'TC 5 dist',
                'TC 6 dist',
            ]

            writer.writerow(row)
            csv_file.close()

    fd = [10, 20, 30, 40, 50]
    n = [5, 7, 9, 11]

    for image in images:
        with open(dir_name + file_name, 'a+') as csv_file:
            writer = csv.writer(csv_file, delimiter=',', quotechar='|', quoting=csv.QUOTE_MINIMAL, lineterminator='\n')
            row = []
            writer.writerow(row)
            row = [
                'correct result',
                '',
                '',
                int(tc1_correct_result[tc1_correct_result['image'] == image]['tc1']),
                int(tc2_correct_result[tc2_correct_result['image'] == image]['tc2']),
                int(tc3_correct_result[tc3_correct_result['image'] == image]['tc3']),
                int(tc4_correct_result[tc4_correct_result['image'] == image]['tc4']),
                int(tc5_correct_result[tc5_correct_result['image'] == image]['tc5']),
                int(tc6_correct_result[tc6_correct_result['image'] == image]['tc6']),
            ]
            writer.writerow(row)
            row = []
            writer.writerow(row)
            csv_file.close()

        for i in fd:
            for j in n:
                with open(dir_name + file_name, 'a+') as csv_file:
                    writer = csv.writer(csv_file, delimiter=',', quotechar='|', quoting=csv.QUOTE_MINIMAL,
                                        lineterminator='\n')
                    row = [
                        image,
                        i,
                        j
                    ]
                    writer.writerow(row)
                    csv_file.close()

    df = pd.read_csv(dir_name + file_name)

    for image in images:
        for i in fd:
            for j in n:
                data_loc_final = data_loc + '%d_fourier_descriptors\\%d_nn\\results.csv' % (i, j)
                data = pd.read_csv(data_loc_final)

                image_results = data.loc[data['Image'] == image]
                tc1_results = image_results[
                    ['Fourier Descriptors', 'Neighbors', 'TC 1 result', 'TC 1 rel', 'TC 1 dist']]
                tc2_results = image_results[
                    ['Fourier Descriptors', 'Neighbors', 'TC 2 result', 'TC 2 rel', 'TC 2 dist']]
                tc3_results = image_results[
                    ['Fourier Descriptors', 'Neighbors', 'TC 3 result', 'TC 3 rel', 'TC 3 dist']]
                tc4_results = image_results[
                    ['Fourier Descriptors', 'Neighbors', 'TC 4 result', 'TC 4 rel', 'TC 4 dist']]
                tc5_results = image_results[
                    ['Fourier Descriptors', 'Neighbors', 'TC 5 result', 'TC 5 rel', 'TC 5 dist']]
                tc6_results = image_results[
                    ['Fourier Descriptors', 'Neighbors', 'TC 6 result', 'TC 6 rel', 'TC 6 dist']]

                results = [tc1_results, tc2_results, tc3_results, tc4_results, tc5_results, tc6_results]

                for ind, result in enumerate(results):
                    df.loc[
                        (df['image'] == image) &
                        (df['Fourier Descriptors'] == i) &
                        (df['Neighbors'] == j), 'TC %d result' % (ind + 1)
                    ] = int(result['TC %d result' % (ind + 1)].iloc[0])

                    df.loc[
                        (df['image'] == image) &
                        (df['Fourier Descriptors'] == i) &
                        (df['Neighbors'] == j), 'TC %d dist' % (ind + 1)
                    ] = result['TC %d dist' % (ind + 1)].iloc[0]

                    df.loc[
                        (df['image'] == image) &
                        (df['Fourier Descriptors'] == i) &
                        (df['Neighbors'] == j), 'TC %d rel' % (ind + 1)
                    ] = result['TC %d rel' % (ind + 1)].iloc[0]

                    correct_result = correct_results_list[ind]
                    correct_result_single_digit = int(
                        correct_result[correct_result['image'] == image]['tc%d' % (ind + 1)])
                    predicted_result_single_digit = int(result['TC %d result' % (ind + 1)].iloc[0])

                    print('image = %s' % image)
                    print('correct result for counter %d: %d' % ((ind + 1), correct_result_single_digit))
                    print('predicted result for counter %d: %d' % ((ind + 1), predicted_result_single_digit))

    df.to_csv(dir_name + file_name, index=False)
    make_excel_sheet(dir_name + file_name, dir_name + xlsx_name)


def make_excel_sheet(data_loc, result_loc):
    """Converts the csv into an excel sheet with formatting to indicate the performance"""

    df = pd.read_csv(data_loc)
    df.to_excel('csv_to_excel.xlsx', index=False)
    workbook = load_workbook(filename='csv_to_excel.xlsx')
    sheet = workbook.active

    os.remove('csv_to_excel.xlsx')

    green_background = PatternFill(bgColor="99d8c9")
    red_background = PatternFill(bgColor="fc9272")

    left_border = Border(
        left=Side(border_style=BORDER_MEDIUM, color='00000000')
    )
    right_border = Border(
        right=Side(border_style=BORDER_MEDIUM, color='00000000')
    )

    up_and_top_border = Border(
        top=Side(border_style=BORDER_MEDIUM, color='00000000'),
        bottom=Side(border_style=BORDER_MEDIUM, color='00000000')
    )

    for i in range(32):
        correct_value_position = 2 + (i * 21)

        for j in range(20):
            cell_pos = 3 + j + (i * 21)
            print("D%d:I%d" % (cell_pos, cell_pos))

            sheet.conditional_formatting.add("$D$%d:$I$%d" % (cell_pos, cell_pos),
                                             CellIsRule(operator='equal', formula=['D$%d' % correct_value_position],
                                                        fill=green_background))

            sheet.conditional_formatting.add("$D$%d:$I$%d" % (cell_pos, cell_pos),
                                             CellIsRule(operator='notEqual', formula=['D$%d' % correct_value_position],
                                                        fill=red_background))

    colors = [Color('d73027'), Color('ffffbf'), Color('1a9850')]
    start = FormatObject(type='num', val=0)
    mid = FormatObject(type='num', val=.5)
    end = FormatObject(type='num', val=1)
    cs3 = ColorScale(cfvo=[start, mid, end], color=colors)
    color_scale_rule = Rule(type='colorScale', colorScale=cs3)

    last_cellnumber = sheet.max_row

    sheet.conditional_formatting.add("J1:J%d" % last_cellnumber, color_scale_rule)
    sheet.conditional_formatting.add("K1:K%d" % last_cellnumber, color_scale_rule)
    sheet.conditional_formatting.add("L1:L%d" % last_cellnumber, color_scale_rule)
    sheet.conditional_formatting.add("M1:M%d" % last_cellnumber, color_scale_rule)
    sheet.conditional_formatting.add("N1:N%d" % last_cellnumber, color_scale_rule)
    sheet.conditional_formatting.add("O1:O%d" % last_cellnumber, color_scale_rule)

    colors = [Color('1a9850'), Color('ffffbf'), Color('d73027')]
    start = FormatObject(type='num', val=0)
    mid = FormatObject(type='num', val=.5)
    end = FormatObject(type='num', val=1.8)
    cs3 = ColorScale(cfvo=[start, mid, end], color=colors)
    color_scale_rule = Rule(type='colorScale', colorScale=cs3)

    sheet.conditional_formatting.add("P1:P%d" % last_cellnumber, color_scale_rule)
    sheet.conditional_formatting.add("Q1:Q%d" % last_cellnumber, color_scale_rule)
    sheet.conditional_formatting.add("R1:R%d" % last_cellnumber, color_scale_rule)
    sheet.conditional_formatting.add("S1:S%d" % last_cellnumber, color_scale_rule)
    sheet.conditional_formatting.add("T1:T%d" % last_cellnumber, color_scale_rule)
    sheet.conditional_formatting.add("U1:U%d" % last_cellnumber, color_scale_rule)

    cols = ['D', 'E', 'F', 'G', 'H', 'I']

    for c in cols:
        for ind, cell in enumerate(sheet['%s' % c]):
            if ind is not 0:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                if c is 'D':
                    cell.border = left_border

    cols = ['J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U']

    for c in cols:
        for ind, cell in enumerate(sheet['%s' % c]):
            if ind is not 0:
                cell.number_format = "0.00"
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                if c is 'J' or c is 'P':
                    cell.border = left_border
                elif c is 'U':
                    cell.border = right_border

    for i in range(32):
        correct_value_position = 2 + (i * 21)

        for cell in sheet[correct_value_position]:
            cell.border = up_and_top_border

        for z in range(6):
            cell = sheet.cell(row=correct_value_position, column=z + 4)
            cell.font = cell.font.copy(bold=True)

    if os.path.exists(result_loc):
        os.remove(result_loc)

    workbook.save(filename=result_loc)


def make_plots_of_accuracy(data_loc, title, column_name, flag_templates, flag_knn):
    """Creates a graph of the classification accuracies"""
    data = pd.read_csv(data_loc)

    import seaborn as sns
    sns.set_context('talk')

    max_col = data[column_name].max()
    min_col = data[column_name].min()

    # Create a grouped bar chart, with job as the x-axis
    # and gender as the variable we're grouping on so there
    # are two bars per job.
    fig, ax = plt.subplots(figsize=(22, 10))
    rcParams.update({'font.size': 26})

    # Our x-axis. We basically just want a list
    # of numbers from zero with a value for each
    # of our jobs.
    x = np.arange(len(data['Neighbors'].unique()))

    plt.plot(x, data.loc[data['Fourier Descriptors'] == 10, column_name],
             label='10 Fourier descriptors', color='#1f78b4', linewidth=2, linestyle=':', marker='o')
    plt.plot(x, data.loc[data['Fourier Descriptors'] == 20, column_name],
             label='20 Fourier descriptors', color='#33a02c', linewidth=2, linestyle=':', marker='o')
    plt.plot(x, data.loc[data['Fourier Descriptors'] == 30, column_name],
             label='30 Fourier descriptors', color='#e31a1c', linewidth=2, linestyle=':', marker='o')
    plt.plot(x, data.loc[data['Fourier Descriptors'] == 40, column_name],
             label='40 Fourier descriptors', color='#ff7f00', linewidth=2, linestyle=':', marker='o')
    plt.plot(x, data.loc[data['Fourier Descriptors'] == 50, column_name],
             label='50 Fourier descriptors', color='#a6cee3', linewidth=2, linestyle=':', marker='o')

    # Fix the x-axes.
    ax.set_xticks(x)

    if max_col <= 1:
        ax.set_yticks(np.arange(0, 1.1, 0.1))
        ax.set_ylim([0, 1])
    else:
        ax.set_yticks(np.arange(min_col - (min_col % 10), 80.5, .5))
        ax.set_ylim([min_col - (min_col % 10), 80])

    ax.set_xticklabels(data['Neighbors'].unique())

    # Add legend.
    ax.legend()

    # Axis styling.
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['left'].set_visible(False)
    ax.spines['bottom'].set_color('#DDDDDD')
    ax.tick_params(bottom=False, left=False)
    ax.set_axisbelow(True)
    ax.yaxis.grid(True, color='#EEEEEE')
    ax.xaxis.grid(False)

    # Add axis and chart labels.
    ax.set_xlabel('Number of Neighbors for KNN classification', labelpad=15)
    ax.set_ylabel('Accuracy in %', labelpad=15)

    fig.tight_layout()

    if not os.path.exists(ROOT_DIR + "\\results\\digit_classification\\accuracy_plots\\"):
        os.makedirs(ROOT_DIR + "\\results\\digit_classification\\accuracy_plots\\structured\\uniform")
        os.makedirs(ROOT_DIR + "\\results\\digit_classification\\accuracy_plots\\structured\\weighted")
        os.makedirs(ROOT_DIR + "\\results\\digit_classification\\accuracy_plots\\random\\uniform")
        os.makedirs(ROOT_DIR + "\\results\\digit_classification\\accuracy_plots\\random\\weighted")

    if flag_templates is 0:
        if flag_knn is 0:
            plt.savefig(
                ROOT_DIR + "\\results\\digit_classification\\accuracy_plots\\structured\\uniform\\%s_plot.pdf" % column_name)
            plt.close()
        else:
            plt.savefig(
                ROOT_DIR + "\\results\\digit_classification\\accuracy_plots\\structured\\weighted\\%s_plot.pdf" % column_name)
            plt.close()
    else:
        if flag_knn is 0:
            plt.savefig(
                ROOT_DIR + "\\results\\digit_classification\\accuracy_plots\\random\\uniform\\%s_plot.pdf" % column_name)
            plt.close()
        else:
            plt.savefig(
                ROOT_DIR + "\\results\\digit_classification\\accuracy_plots\\random\\weighted\\%s_plot.pdf" % column_name)
            plt.close()


def make_plots_line(data_loc, title, column_name_1, column_name_2, flag_templates, flag_knn):
    """Creates a plot comparing the performance of  the correct and incoreect classifications per classification run"""

    data = pd.read_csv(data_loc)

    import seaborn as sns
    sns.set_context('talk')

    max_col = data[[column_name_1, column_name_2]].max().max()
    min_col = data[[column_name_1, column_name_2]].min().min()

    rcParams.update({'font.size': 22})

    # Create a grouped bar chart, with job as the x-axis
    # and gender as the variable we're grouping on so there
    # are two bars per job.
    fig, (ax1, ax2, ax3) = plt.subplots(nrows=1, ncols=3, sharex=True, sharey=True, figsize=(30, 15),
                                        gridspec_kw={'width_ratios': [3, 3, 1]})
    # ax.plot([1, 2])

    # Our x-axis. We basically just want a list
    # of numbers from zero with a value for each
    # of our jobs.
    x = np.arange(len(data['Neighbors'].unique()))

    # Define bar width. We need this to offset the second bar.

    ax1.plot(x, data.loc[data['Fourier Descriptors'] == 10, column_name_1],
             label='10 Fourier descriptors', color='#1f78b4', linewidth=2.5, linestyle=':', marker='o')
    ax1.plot(x, data.loc[data['Fourier Descriptors'] == 20, column_name_1],
             label='20 Fourier descriptors', color='#33a02c', linewidth=2.5, linestyle=':', marker='o')
    ax1.plot(x, data.loc[data['Fourier Descriptors'] == 30, column_name_1],
             label='30 Fourier descriptors', color='#e31a1c', linewidth=2.5, linestyle=':', marker='o')
    ax1.plot(x, data.loc[data['Fourier Descriptors'] == 40, column_name_1],
             label='40 Fourier descriptors', color='#ff7f00', linewidth=2.5, linestyle=':', marker='o')
    ax1.plot(x, data.loc[data['Fourier Descriptors'] == 50, column_name_1],
             label='50 Fourier descriptors', color='#a6cee3', linewidth=2.5, linestyle=':', marker='o')

    ax2.plot(x, data.loc[data['Fourier Descriptors'] == 10, column_name_2],
             label='10 Fourier descriptors', color='#1f78b4', linewidth=2.5, linestyle=':', marker='o')
    ax2.plot(x, data.loc[data['Fourier Descriptors'] == 20, column_name_2],
             label='20 Fourier descriptors', color='#33a02c', linewidth=2.5, linestyle=':', marker='o')
    ax2.plot(x, data.loc[data['Fourier Descriptors'] == 30, column_name_2],
             label='30 Fourier descriptors', color='#e31a1c', linewidth=2.5, linestyle=':', marker='o')
    ax2.plot(x, data.loc[data['Fourier Descriptors'] == 40, column_name_2],
             label='40 Fourier descriptors', color='#ff7f00', linewidth=2.5, linestyle=':', marker='o')
    ax2.plot(x, data.loc[data['Fourier Descriptors'] == 50, column_name_2],
             label='50 Fourier descriptors', color='#a6cee3', linewidth=2.5, linestyle=':', marker='o')

    ax3.plot(x, data.loc[data['Fourier Descriptors'] == 10, column_name_2],
             label='10 Fourier descriptors', color='#1f78b4', linewidth=2.5, linestyle=':', marker='o')
    ax3.plot(x, data.loc[data['Fourier Descriptors'] == 20, column_name_2],
             label='20 Fourier descriptors', color='#33a02c', linewidth=2.5, linestyle=':', marker='o')
    ax3.plot(x, data.loc[data['Fourier Descriptors'] == 30, column_name_2],
             label='30 Fourier descriptors', color='#e31a1c', linewidth=2.5, linestyle=':', marker='o')
    ax3.plot(x, data.loc[data['Fourier Descriptors'] == 40, column_name_2],
             label='40 Fourier descriptors', color='#ff7f00', linewidth=2.5, linestyle=':', marker='o')
    ax3.plot(x, data.loc[data['Fourier Descriptors'] == 50, column_name_2],
             label='50 Fourier descriptors', color='#a6cee3', linewidth=2.5, linestyle=':', marker='o')

    l = ax3.legend(loc="center", fontsize='large')
    # ax3.set_visible(False)

    for line in ax3.lines:
        line.set_visible(False)

    ax3.spines['top'].set_visible(False)
    ax3.spines['right'].set_visible(False)
    ax3.spines['left'].set_visible(False)
    ax3.spines['bottom'].set_visible(False)
    ax3.axis('off')

    # Fix the x-axes.
    ax1.set_xticks(x)
    ax2.set_xticks(x)

    if max_col <= 1:
        ax1.set_yticks(np.arange(0, 1.1, 0.1))
        ax2.set_yticks(np.arange(0, 1.1, 0.1))
        ax1.set_ylim([min_col - (min_col % 0.1), max_col + (.1 - max_col % .1)])
        ax2.set_ylim([min_col - (min_col % 0.1), max_col + (.1 - max_col % .1)])
    else:
        ax1.set_yticks(np.arange(min_col - (min_col % 10), 102.5, 2.5))
        ax2.set_yticks(np.arange(min_col - (min_col % 10), 102.5, 2.5))
        ax1.set_ylim([min_col - (min_col % 10), 100])
        ax2.set_ylim([min_col - (min_col % 10), 100])

    labels = data['Neighbors'].unique()

    ax1.set_xticklabels(labels)
    ax2.set_xticklabels(labels)

    # Axis styling.
    ax1.spines['top'].set_visible(False)
    ax2.spines['top'].set_visible(False)
    ax1.spines['right'].set_visible(False)
    ax2.spines['right'].set_visible(False)
    ax1.spines['left'].set_visible(False)
    ax2.spines['left'].set_visible(False)
    ax1.spines['bottom'].set_color('#DDDDDD')
    ax2.spines['bottom'].set_color('#DDDDDD')
    ax1.tick_params(bottom=False, left=False)
    ax2.tick_params(bottom=False, left=False)
    ax1.set_axisbelow(True)
    ax2.set_axisbelow(True)
    ax1.yaxis.grid(True, color='#EEEEEE')
    ax2.yaxis.grid(True, color='#EEEEEE')
    ax1.xaxis.grid(False)
    ax2.xaxis.grid(False)

    ax1.set_title("Correct", pad=15)
    ax2.set_title("Incorrect", pad=15)

    # Add axis and chart labels.
    ax1.set_xlabel('Number of Neighbors', labelpad=15)
    ax2.set_xlabel('Number of Neighbors', labelpad=15)

    if column_name_1 == 'Average Reliability Correct' or column_name_1 == 'Standard Deviation Reliability Correct':
        ax1.set_ylabel('Reliability Score', labelpad=15)
    else:
        ax1.set_ylabel('Euclidean Distance', labelpad=15)

    fig.suptitle(title)
    # fig.tight_layout()

    if not os.path.exists(ROOT_DIR + "\\results\\digit_classification\\accuracy_plots\\"):
        os.makedirs(ROOT_DIR + "\\results\\digit_classification\\accuracy_plots\\structured\\uniform")
        os.makedirs(ROOT_DIR + "\\results\\digit_classification\\accuracy_plots\\structured\\weighted")
        os.makedirs(ROOT_DIR + "\\results\\digit_classification\\accuracy_plots\\random\\uniform")
        os.makedirs(ROOT_DIR + "\\results\\digit_classification\\accuracy_plots\\random\\weighted")

    if flag_templates is 0:
        if flag_knn is 0:
            plt.savefig(
                ROOT_DIR + "\\results\\digit_classification\\accuracy_plots\\structured\\uniform\\%s_plot vs %s_plot.pdf" % (
                    column_name_1, column_name_2))
            plt.close()
        else:
            plt.savefig(
                ROOT_DIR + "\\results\\digit_classification\\accuracy_plots\\structured\\weighted\\%s_plot vs %s_plot.pdf" % (
                    column_name_1, column_name_2))
            plt.close()
    else:
        if flag_knn is 0:
            plt.savefig(
                ROOT_DIR + "\\results\\digit_classification\\accuracy_plots\\random\\uniform\\%s_plot vs %s_plot.pdf" % (
                    column_name_1, column_name_2))
            plt.close()
        else:
            plt.savefig(
                ROOT_DIR + "\\results\\digit_classification\\accuracy_plots\\random\\weighted\\%s_plot vs %s_plot.pdf" % (
                    column_name_1, column_name_2))
            plt.close()


def make_plots_bar(data_loc, title, column_name_1, column_name_2, flag_templates, flag_knn):
    """Creates a bar chart comparing the performance of  the correct and incoreect classifications
    per classification run"""

    data = pd.read_csv(data_loc)

    import seaborn as sns
    sns.set_context('talk')

    rcParams.update({'font.size': 22})

    max_col = data[[column_name_1, column_name_2]].max().max()
    min_col = data[[column_name_1, column_name_2]].min().min()

    # Create a grouped bar chart, with job as the x-axis
    # and gender as the variable we're grouping on so there
    # are two bars per job.
    fig, (ax1, ax2, ax3) = plt.subplots(nrows=1, ncols=3, sharex=True, sharey=True, figsize=(30, 15),
                                        gridspec_kw={'width_ratios': [4, 4, 1]})
    # ax.plot([1, 2])

    # Our x-axis. We basically just want a list
    # of numbers from zero with a value for each
    # of our jobs.
    x = np.arange(len(data['Neighbors'].unique()))

    # Define bar width. We need this to offset the second bar.
    bar_width = 0.12

    b1_1 = ax1.bar(x,
                   data.loc[data['Fourier Descriptors'] == 10, column_name_1],
                   width=bar_width, label='10 Fourier descriptors', color='#fbb4ae')

    # yerr=data.loc[data['Fourier Descriptors'] == 10, std_1], capsize=5, ecolor='#e31a1c',
    b1_2 = ax1.bar(x + bar_width,
                   data.loc[data['Fourier Descriptors'] == 20, column_name_1],
                   width=bar_width, label='20 Fourier descriptors', color='#b3cde3')

    # yerr=data.loc[data['Fourier Descriptors'] == 20, std_1], capsize=5, ecolor='#1f78b4',

    b1_3 = ax1.bar(x + 2 * bar_width,
                   data.loc[data['Fourier Descriptors'] == 30, column_name_1],
                   width=bar_width, label='30 Fourier descriptors', color='#ccebc5')

    # yerr=data.loc[data['Fourier Descriptors'] == 30, std_1], capsize=5, ecolor='#33a02c',

    b1_4 = ax1.bar(x + 3 * bar_width,
                   data.loc[data['Fourier Descriptors'] == 40, column_name_1],
                   width=bar_width, label='40 Fourier descriptors', color='#decbe4')

    # yerr=data.loc[data['Fourier Descriptors'] == 40, std_1], capsize=5, ecolor='#6a3d9a',

    b1_5 = ax1.bar(x + 4 * bar_width,
                   data.loc[data['Fourier Descriptors'] == 50, column_name_1],
                   width=bar_width, label='50 Fourier descriptors', color='#fed9a6')

    # yerr=data.loc[data['Fourier Descriptors'] == 50, std_1], capsize=5, ecolor='#ff7f00',

    b2_1 = ax2.bar(x,
                   data.loc[data['Fourier Descriptors'] == 10, column_name_2],
                   width=bar_width, label='10 Fourier descriptors', color='#fbb4ae')

    # yerr=data.loc[data['Fourier Descriptors'] == 10, std_2], capsize=5, ecolor='#e31a1c',
    b2_2 = ax2.bar(x + bar_width,
                   data.loc[data['Fourier Descriptors'] == 20, column_name_2],
                   width=bar_width, label='20 Fourier descriptors', color='#b3cde3')

    # yerr=data.loc[data['Fourier Descriptors'] == 20, std_2], capsize=5, ecolor='#1f78b4',

    b2_3 = ax2.bar(x + 2 * bar_width,
                   data.loc[data['Fourier Descriptors'] == 30, column_name_2],
                   width=bar_width, label='30 Fourier descriptors', color='#ccebc5')

    # yerr=data.loc[data['Fourier Descriptors'] == 30, std_2], capsize=5, ecolor='#33a02c',

    b2_4 = ax2.bar(x + 3 * bar_width,
                   data.loc[data['Fourier Descriptors'] == 40, column_name_2],
                   width=bar_width, label='40 Fourier descriptors', color='#decbe4')

    # yerr=data.loc[data['Fourier Descriptors'] == 40, std_2], capsize=5, ecolor='#6a3d9a',

    b2_5 = ax2.bar(x + 4 * bar_width,
                   data.loc[data['Fourier Descriptors'] == 50, column_name_2],
                   width=bar_width, label='50 Fourier descriptors', color='#fed9a6')

    # yerr=data.loc[data['Fourier Descriptors'] == 50, std_2], capsize=5, ecolor='#ff7f00',

    b3_1 = ax3.bar(x,
                   data.loc[data['Fourier Descriptors'] == 10, column_name_1],
                   width=bar_width, label='10 Fourier descriptors', color='#fbb4ae')
    b3_2 = ax3.bar(x + bar_width,
                   data.loc[data['Fourier Descriptors'] == 20, column_name_1],
                   width=bar_width, label='20 Fourier descriptors', color='#b3cde3')

    b3_3 = ax3.bar(x + 2 * bar_width,
                   data.loc[data['Fourier Descriptors'] == 30, column_name_1],
                   width=bar_width, label='30 Fourier descriptors', color='#ccebc5')

    b3_4 = ax3.bar(x + 3 * bar_width,
                   data.loc[data['Fourier Descriptors'] == 40, column_name_1],
                   width=bar_width, label='40 Fourier descriptors', color='#decbe4')

    b3_5 = ax3.bar(x + 4 * bar_width,
                   data.loc[data['Fourier Descriptors'] == 50, column_name_1],
                   width=bar_width, label='50 Fourier descriptors', color='#fed9a6')

    l = ax3.legend(loc="center", fontsize='large')
    # ax3.set_visible(False)

    for bar in ax3.patches:
        bar.set_visible(False)

    ax3.spines['top'].set_visible(False)
    ax3.spines['right'].set_visible(False)
    ax3.spines['left'].set_visible(False)
    ax3.spines['bottom'].set_visible(False)
    ax3.axis('off')

    # Fix the x-axes.
    ax1.set_xticks(x + bar_width * 4 / 2)
    ax2.set_xticks(x + bar_width * 4 / 2)

    if max_col <= 1:
        ax1.set_yticks(np.arange(0, 1.1, 0.1))
        ax2.set_yticks(np.arange(0, 1.1, 0.1))

        if max_col + .1 + (.1 - max_col % .1) > 1:
            max_lim = 1
        else:
            max_lim = max_col + .1 + (.1 - max_col % .1)
        ax1.set_ylim(min_col - .1 - (min_col % .1), max_lim)
        ax2.set_ylim(min_col - .1 - (min_col % .1), max_lim)
    else:
        ax1.set_yticks(np.arange(min_col - (min_col % 10), 102.5, 2.5))
        ax2.set_yticks(np.arange(min_col - (min_col % 10), 102.5, 2.5))
        ax1.set_ylim([min_col - (min_col % 10), 100])
        ax2.set_ylim([min_col - (min_col % 10), 100])

    labels = data['Neighbors'].unique()

    ax1.set_xticklabels(labels)
    ax2.set_xticklabels(labels)

    # Axis styling.
    ax1.spines['top'].set_visible(False)
    ax2.spines['top'].set_visible(False)
    ax1.spines['right'].set_visible(False)
    ax2.spines['right'].set_visible(False)
    ax1.spines['left'].set_visible(False)
    ax2.spines['left'].set_visible(False)
    ax1.spines['bottom'].set_color('#DDDDDD')
    ax2.spines['bottom'].set_color('#DDDDDD')
    ax1.tick_params(bottom=False, left=False)
    ax2.tick_params(bottom=False, left=False)
    ax1.set_axisbelow(True)
    ax2.set_axisbelow(True)
    ax1.yaxis.grid(True, color='#EEEEEE')
    ax2.yaxis.grid(True, color='#EEEEEE')
    ax1.xaxis.grid(False)
    ax2.xaxis.grid(False)
    ax1.set_title("Correct", pad=15)
    ax2.set_title("Incorrect", pad=15)

    # Add axis and chart labels.
    ax1.set_xlabel('Number of Neighbors', labelpad=15)
    ax2.set_xlabel('Number of Neighbors', labelpad=15)

    if column_name_1 == 'Average Reliability Correct' or column_name_1 == 'Standard Deviation Reliability Correct':
        ax1.set_ylabel('Reliability Score', labelpad=15)
    else:
        ax1.set_ylabel('Euclidean Distance', labelpad=15)

    fig.suptitle(title)
    # fig.tight_layout()

    if not os.path.exists(ROOT_DIR + "\\results\\digit_classification\\accuracy_plots\\"):
        os.makedirs(ROOT_DIR + "\\results\\digit_classification\\accuracy_plots\\structured\\uniform")
        os.makedirs(ROOT_DIR + "\\results\\digit_classification\\accuracy_plots\\structured\\weighted")
        os.makedirs(ROOT_DIR + "\\results\\digit_classification\\accuracy_plots\\random\\uniform")
        os.makedirs(ROOT_DIR + "\\results\\digit_classification\\accuracy_plots\\random\\weighted")

    if flag_templates is 0:
        if flag_knn is 0:
            plt.savefig(
                ROOT_DIR + "\\results\\digit_classification\\accuracy_plots\\structured\\uniform\\%s_plot vs %s_barchart.pdf" % (
                    column_name_1, column_name_2))
            plt.close()
        else:
            plt.savefig(
                ROOT_DIR + "\\results\\digit_classification\\accuracy_plots\\structured\\weighted\\%s_plot vs %s_barchart.pdf" % (
                    column_name_1, column_name_2))
            plt.close()
    else:
        if flag_knn is 0:
            plt.savefig(
                ROOT_DIR + "\\results\\digit_classification\\accuracy_plots\\random\\uniform\\%s_plot vs %s_barchart.pdf" % (
                    column_name_1, column_name_2))
            plt.close()
        else:
            plt.savefig(
                ROOT_DIR + "\\results\\digit_classification\\accuracy_plots\\random\\weighted\\%s_plot vs %s_barchart.pdf" % (
                    column_name_1, column_name_2))
            plt.close()


if __name__ == '__main__':
    flags_templates = [1]  # 0 is structured templates and 1 is random templates
    flags_knn = [0, 1]  # 0 is uniform and 1 is weighted

    dir_name = ROOT_DIR + '\\results\\digit_classification\\accuracy_extended\\'

    for i in flags_templates:
        if i is 0:
            data_loc = ROOT_DIR + "\\results\\digit_classification\\detailed_results_separated\\digit_classification_extendedstructured"
            name = 'classification_accuracy_structured_'
        elif i is 1:
            data_loc = ROOT_DIR + "\\results\\digit_classification\\detailed_results_separated\\digit_classification_extendedrandom"
            name = 'classification_accuracy_random_'
        for j in flags_knn:
            if j is 0:
                data_loc_final = data_loc + "_uniform\\"
                file_name = name + 'uniform.csv'
                make_csv(data_loc_final, dir_name, file_name, flag_templates=i, flag_knn=j)
                column_names = ["Accuracy",
                                ("Average Reliability Correct", "Average Reliability Incorrect"),
                                (
                                    "Confidence Interval Reliability Correct",
                                    "Confidence Interval Reliability Incorrect"),
                                ("Average Euclidean Distance Correct", "Average Euclidean Distance Incorrect"),
                                ("Confidence Interval Euclidean Distance Correct",
                                 "Confidence Interval Euclidean Distance Incorrect")
                                ]

                for index, column_name in enumerate(column_names):
                    if type(column_name) is not tuple:
                        title = "%s for Uniform Weighted Digit classification" % column_name
                        # make_plots_of_accuracy(dir_name + file_name, title, column_name, i, j)
                    else:
                        c1, c2 = column_name
                        title = "%s vs %s for Uniform Weighted Digit classification" % (c1, c2)
                        if c1 == "Average Reliability Correct" or c1 == "Average Euclidean Distance Correct":
                            std1, std2 = column_names[index + 1]
                            # make_plots_bar(dir_name + file_name, title, c1, c2, std1, std2, i, j)
                        # make_plots_line(dir_name + file_name, title, c1, c2, i, j)

            elif j is 1:
                data_loc_final = data_loc + "_weighted\\"
                file_name = name + 'weighted.csv'
                make_csv(data_loc_final, dir_name, file_name, flag_templates=i, flag_knn=j)

                column_names = ["Accuracy",
                                ("Average Reliability Correct", "Average Reliability Incorrect"),
                                (
                                    "Confidence Interval Reliability Correct",
                                    "Confidence Interval Reliability Incorrect"),
                                ("Average Euclidean Distance Correct", "Average Euclidean Distance Incorrect"),
                                ("Confidence Interval Euclidean Distance Correct",
                                 "Confidence Interval Euclidean Distance Incorrect")
                                ]

                for index, column_name in enumerate(column_names):
                    if type(column_name) is not tuple:
                        title = "%s for Distance Weighted Digit classification" % column_name
                        # make_plots_of_accuracy(dir_name + file_name, title, column_name, i, j)
                    else:
                        c1, c2 = column_name
                        title = "%s vs %s for Distance Weighted Digit classification" % (c1, c2)
                        if c1 == "Average Reliability Correct" or c1 == "Average Euclidean Distance Correct":
                            std1, std2 = column_names[index + 1]
                            # make_plots_bar(dir_name + file_name, title, c1, c2, std1, std2, i, j)
                        # make_plots_line(dir_name + file_name, title, c1, c2, i, j)

    dir_name = ROOT_DIR + '\\results\\digit_classification\\detailed_results_extended\\'

    for i in flags_templates:
        if i is 0:
            data_loc = ROOT_DIR + "\\results\\digit_classification\\detailed_results_separated\\digit_classification_extendedstructured"
            name = 'classification_details_structured_'
        elif i is 1:
            data_loc = ROOT_DIR + "\\results\\digit_classification\\detailed_results_separated\\digit_classification_extendedrandom"
            name = 'classification_details_random_'
        for j in flags_knn:
            if j is 0:
                data_loc_final = data_loc + "_uniform\\"
                file_name = name + 'uniform.csv'
                xlsx_name = name + 'uniform.xlsx'
                make_csv_detailed(data_loc_final, dir_name, file_name, xlsx_name)
            elif j is 1:
                data_loc_final = data_loc + "_weighted\\"
                file_name = name + 'weighted.csv'
                xlsx_name = name + 'weighted.xlsx'
                make_csv_detailed(data_loc_final, dir_name, file_name, xlsx_name)
